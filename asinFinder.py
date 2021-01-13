import openpyxl
from pathlib import Path
import glob


all_listing_report = Path(Path.home(),'Dropbox','LISTING','NewListings','allListingReport.xlsx')
master_sheet = Path(Path.home(),'Dropbox','masterList.xlsx')
dest_name = "Master_List_With_ASIN.xlsx"

if __name__ == "__main__":
    #open both the masterList and the allListingReport
    val = input("Enter the date the all listing report was downloaded (mm/dd/yyyy): ")
    month = val[:2]
    day = val[3:5]
    year = val[6:]

    listingReportSheetName = 'All+Listings+Report+' + month + '-' + day + '-' + year

    ms = openpyxl.load_workbook(master_sheet)['Sheet']
    alr = openpyxl.load_workbook(all_listing_report)[listingReportSheetName]

    fw = openpyxl.Workbook()
    fs = fw['Sheet']

    max = ms.max_row
    alrSKU = alr['D']

    fs.cell(1,1).value = 'SKU'
    fs.cell(1,2).value = 'ASIN'
    fs.cell(1,3).value = 'STYLE SHEET'
    for x in range(1,max+1):

        sku = ms.cell(x,1).value
        for counter, sku_1 in enumerate(alrSKU,start=1):
            if sku_1.value == sku:
                print(f"Found ASIN {x}/{max}")
                final = fs.max_row + 1
                fs.cell(final,1).value = sku
                fs.cell(final,2).value = alr.cell(counter,17).value
                fs.cell(final,3).value = ms.cell(x,3).value
                fs.cell(final,4).value = ms.cell(x,4).value
                fs.cell(final,5).value = ms.cell(x,5).value

                break
    fw.save(dest_name)
    fw.close()
    #ms.close()
    #alr.close()
    #fs.close()
