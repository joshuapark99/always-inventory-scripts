import openpyxl
from pathlib import Path
import typing
import glob


master_sheet = Path('LISTING','NewListings','masterList.xlsx')
# get all Excel files from Dropbox folder and store in a list
xlsx_files = [path for path in Path('LISTING').rglob('*.xlsx')]
xlsm_files = [path for path in Path('LISTING').rglob('*.xlsm')]
# create a workbook object for each file in xlsx_files list

#wbObjects = [openpyxl.load_workbook(wb,read_only = True) for wb in xlsx_files]
# max of 250 open files

def getMultipleSKU(filePaths, startIndex = 0,numberToOpen = 50):

    sku = []
    for path in filePaths[startIndex:(startIndex+numberToOpen)]:
        if str(path)[0:9] == "LISTING/~":
            continue
        print(f'working on {path}')
        wb = openpyxl.load_workbook(path,read_only=True)

        skuList = getSKUFromSheet(wb)

        for skuNumber in skuList:
            if(skuNumber is not None):
                sku.append((skuNumber,str(path)))
        wb.close()
    print(sku)
    return sku

def getSKUFromSheet(wb):

    SKUList = []
    style_sheet = ''
    sheets = wb.sheetnames;
    if('Template' in sheets):
        style_sheet = wb['Template']
    else:
        for sheet in sheets:
            temp = wb[sheet].cell(1,1).value
            if type(temp) == type(' ') and temp[:12] == 'TemplateType':
                style_sheet = wb[sheet]
                break
        if style_sheet == '':
            return []

    highest_row = style_sheet.max_row
    values = []
    for row in range(1,4):
        for col in range(1,5):
            if style_sheet.cell(row,col).value == "item_sku":
                #SKUList = getValueColumn(style_sheet,row,col,max)
                for counter in range(4,highest_row+1):
                    SKUList.append(style_sheet.cell(counter,col).value)

    return SKUList

def getValueColumn(style_sheet, row, col, max):
    #print('got there')
    values = []
    for counter in range(4,max):
        values.append(style_sheet.cell(counter, col).value)
    return values

def inputToMaster(ms,skuList):
    #wb = openpyxl.load_workbook(master_sheet)
    for count,x in enumerate(skuList):
        inputRow = ms['SKU'].max_row+1
        ms['SKU'].cell(inputRow, 1).value = x[0]
        ms['SKU'].cell(inputRow, 3).value = x[1]
        print(f"writing progress {count}/{len(skuList)}")
    #wb.save('masterList.xlsx')
    #wb.close()

#def inputToSheet(skuList):
if __name__ == "__main__":
    print('hi')
    skuSheet = []
    ms = openpyxl.load_workbook(master_sheet)
    for x in range(int(len(xlsx_files)/50)+1):
        print(f'working on {x+1}/{int(len(xlsx_files)/50+1)}')
        skuSheet = getMultipleSKU(xlsx_files, x*50)
        inputToMaster(ms,skuSheet)
    for x in range(int(len(xlsm_files)/50)+1):
        print(f'working on {x+1}/{int(len(xlsm_files)/50)+1}')
        skuSheet = getMultipleSKU(xlsm_files, x*50)
        inputToMaster(ms,skuSheet)
    ms.save('masterList.xlsx')
    ms.close()
    #inputToMaster(getMultipleSKU(xlsx_files, 0,5))
