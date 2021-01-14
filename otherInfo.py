import openpyxl
from pathlib import Path
import glob
from collections import defaultdict
import multiprocessing as mp
import time



# takes too long

masterSheet = Path("Master_List_With_ASIN.xlsx")

dest = "Master_List_With_All_Info.xlsx"

workbooks = defaultdict(lambda: "not yet")

other_info = ['item_dimensions_unit_of_measure', 'website_shipping_weight', 'website_shipping_weight_unit_of_measure', 'item_length', 'item_width', 'item_height','fabric_type']
knownColumns = {'item_dimensions_unit_of_measure': None, 'website_shipping_weight': None, 'website_shipping_weight_unit_of_measure':None, 'item_length':None, 'item_width':None, 'item_height':None, 'fabric_type':None}
lastUsedWBPath = None
wb = None
style_sheet = None
first_run = 0


def getInfo(sku,path,row):
    #print(path)
    global lastUsedWBPath
    global wb
    global style_sheet
    global knownColumns
    sameWB = True
    object = defaultdict(lambda: "not here")
    print(f"{lastUsedWBPath} - {path}")
    if(lastUsedWBPath == None or lastUsedWBPath != path):

        print("NEW PATH")
        sameWB = False
        if(wb != None):
            wb.close()
        wb = openpyxl.load_workbook(Path(path),read_only=True)
        lastUsedWBPath = path
        style_sheet = ''
        sheets = wb.sheetnames;
        #object = defaultdict(lambda: "not here")
        if('Template' in sheets):
            style_sheet = wb['Template']
        else:
            for sheet in sheets:
                temp = wb[sheet].cell(1,1).value
                if type(temp) == type(' ') and temp[:12] == 'TemplateType':
                    style_sheet = wb[sheet]
                    break
    #wb = openpyxl.load_workbook(Path(path),read_only=True)

    # if its the same workbook, just use the column variables that are saved
    if(sameWB):
        for key in knownColumns:
            temp = knownColumns[key]
            if(temp is not None):
                object[key] = style_sheet.cell(row,temp).value
    else:
        counter = 0
        knownColumns = dict.fromkeys(other_info)
        for col2 in range(1,style_sheet.max_column+1):
            temp = style_sheet.cell(3,col2).value
            if temp in other_info:
                counter += 1
                knownColumns[temp] = col2
                object[temp] = style_sheet.cell(row,col2).value
            if counter == len(other_info):
                break
    #wb.close()
    return object

if __name__ == "__main__":
    #global lastUsedWBPath
    start_time = time.time()
    m_wb = openpyxl.load_workbook(masterSheet,read_only=True)
    new_wb = openpyxl.Workbook()
    pool = mp.Pool(mp.cpu_count())
    path = ''
    read_sheet = m_wb['Sheet']


    write_sheet = new_wb['Sheet']

    for counter in range(2,read_sheet.max_row+1):
        sku = read_sheet.cell(counter,1).value
        print(path)
        if(counter%50 == 0):
            for tempV in range(10):
                print(time.time()-start_time)
        #if(path != lastUsedWBPath):
        path = read_sheet.cell(counter,3).value
        row = read_sheet.cell(counter,4).value
        #col = read_sheet.cell(counter,5).value
        object = getInfo(sku, path, row)
        """write_sheet.cell(counter,1).value = sku
        write_sheet.cell(counter,2).value = read_sheet.cell(counter,2).value
        write_sheet.cell(counter,3).value = path
        for x,y in enumerate(other_info,start=6):
            if object[y] != "not here":
                write_sheet.cell(counter,x).value = object[y]"""
        otherInfoList = [object[k] for k in other_info]
        write_sheet.append([sku,read_sheet.cell(counter,2).value, path] + otherInfoList)
        print(f'done with item {counter}/{read_sheet.max_row}')

    new_wb.save(dest)
    new_wb.close()
    m_wb.close()
